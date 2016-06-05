import openpyxl

profs = []

class Professor(object):
    def __init__(self, ID, name):
        self.ID = ID
        self.name = name

def load_profs():
    global profs
    wb = openpyxl.load_workbook('Professors.xlsx')
    sh = wb['Professors']

    starting_row = 0
    starting_col = 0

    while sh.cell(row=starting_row+1, column=starting_col+1).value:
        profs.append(Professor(ID = starting_row, name = sh.cell(row=starting_row+1, column=starting_col+1).value))
        starting_row += 1