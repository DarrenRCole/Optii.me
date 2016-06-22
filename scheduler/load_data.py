from models import *
from openpyxl import *

NUM_DAILY_TIMESLOTS = 20
NUM_DAYS_IN_WEEK = 5
PROF_H_CONSTRAINT_CODE = 2
PROF_S_CONSTRAINT_CODE = 1
TIMESLOTS = ['8:00', '8:30', '9:00', '9:30','10:00', '10:30','11:00', '11:30','12:00', '12:30','13:00', '13:30','14:00', '14:30','15:00', '15:30','16:00', '16:30','17:00', '17:30']
DAYS = ['m', 't', 'w', 'r', 'f']


def create_profs():
	return false
	#TO BE WRITTEN

def create_rooms():
	return false
	#TO BE WRITTEN

def create_courses():
	return false
	#TO BE WRITTEN

def create_sections():
	wb = load_workbook('testdata.xlsx')
	sheet = wb.get_sheet_by_name('sections')

	row = 2
	course_id_offset = 3
	prof_id_offset = 4

	while sheet.cell(row=row, column=1).value is not None:
		prof_id = sheet.cell(row=row, column=1).value + prof_id_offset
		course_id = sheet.cell(row=row, column=2).value + course_id_offset
		
		prof = Professor.objects.filter(id=prof_id)
		course = Course.objects.filter(id=course_id)

		Section.objects.create(professor=prof[0], course=course[0])
		row += 1

def create_offerings():
	wb = load_workbook('testdata.xlsx')
	sheet = wb.get_sheet_by_name('offerings')

	row = 2
	section_id_offset = 15

	while sheet.cell(row=row, column=1).value is not None:
		section_id = sheet.cell(row=row, column=3).value + section_id_offset
		
		duration = sheet.cell(row=row, column=1).value
		capacity = sheet.cell(row=row, column=2).value
		section = Section.objects.filter(id=section_id)


		Offering.objects.create(duration=duration, capacity=capacity, section=section[0])
		row += 1

def create_unavailabilities():
	wb = load_workbook('testdata.xlsx')
	sheet = wb.get_sheet_by_name('prof unavailabilities')

	starting_column = 2
	row = 3

	while sheet.cell(row=row, column=1).value is not None:
		prof_id = sheet.cell(row=row, column=1).value
		prof = Professor.objects.filter(id=prof_id)
		prof = prof[0]

		for col in range (NUM_DAILY_TIMESLOTS*NUM_DAYS_IN_WEEK):
			unavail_value = sheet.cell(row=row, column=col+starting_column).value
			print unavail_value
			if ((unavail_value is not None) and (unavail_value != 0)):
				day = DAYS[sheet.cell(row=2, column=col+starting_column).value-1]
				start_time = TIMESLOTS[col%NUM_DAILY_TIMESLOTS]
				end_time =  TIMESLOTS[(col+1)%NUM_DAILY_TIMESLOTS]
				preference_level = unavail_value
				
				ProfessorUnavailability.objects.create(day=day, start_time=start_time, end_time=end_time, preference_level=preference_level, professor=prof)

		row += 1
