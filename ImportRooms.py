import openpyxl

rooms = []

class Room(object):
    def __init__(self, ID, name, capacity):
        self.ID = ID
        self.name = name
        self.capacity = capacity

def load_rooms():
    global rooms

    wb = openpyxl.load_workbook('Rooms.xlsx')
    sh = wb['Rooms']

    row = 1
    column = 1

    while sh.cell(row=row, column=column).value:
        rooms.append(Room(ID = row-1,
                          name = sh.cell(row=row, column=column).value,
                          capacity = int(sh.cell(row=row, column=column+1).value)))
        row += 1