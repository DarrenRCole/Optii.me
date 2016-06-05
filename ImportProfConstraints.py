import openpyxl
import LoadTimeslots
import ImportProfs

prof_constraints = []

def load_prof_constraints():
    global prof_constraints
    prof_constraints = [[0 for x in range(LoadTimeslots.num_timeslots)] for x in range(len(ImportProfs.profs))]
    wb = openpyxl.load_workbook('Professors.xlsx')
    sh = wb['Professors']
    starting_row = 1
    starting_col = 2

    for y in range(LoadTimeslots.num_timeslots):
        for x in range(len(ImportProfs.profs)):
            if(sh.cell(row=starting_row+x, column=starting_col+y).value == 'h'):
                prof_constraints[x][y] = 2
            elif(sh.cell(row=starting_row+x, column=starting_col+y).value == 's'):
                prof_constraints[x][y] = 1
            else:
                prof_constraints[x][y] = 0