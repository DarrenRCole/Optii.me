import openpyxl

course_sections = []

class CourseSection(object):
    def __init__(self, ID, course, section, capacity, room_timeslot_ID, prof_ID):
        self.ID = ID
        self.course = course
        self.section = section
        self.capacity = capacity
        self.room_timeslot_ID = room_timeslot_ID
        self.prof_ID = prof_ID

def load_course_sections():
    global course_sections
    wb = openpyxl.load_workbook('Course Sections.xlsx')
    sh = wb['Course Sections']

    starting_row = 1
    row_iter = 0
    starting_col = 1

    while sh.cell(row=starting_row+row_iter, column=starting_col).value:
        course_sections.append(CourseSection(ID = row_iter,
                                             course = sh.cell(row=starting_row + row_iter, column=starting_col).value,
                                             section = sh.cell(row=starting_row + row_iter, column=starting_col+1).value,
                                             capacity = int(sh.cell(row=starting_row + row_iter, column=starting_col+2).value),
                                             room_timeslot_ID = False,
                                             prof_ID = sh.cell(row=starting_row + row_iter, column=starting_col+3).value))
        row_iter += 1